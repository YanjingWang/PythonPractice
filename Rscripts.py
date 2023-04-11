import pandas as pd
import pyodbc
import numpy as np
from datetime import datetime
import os
import re
import pandas as pd
from openpyxl import load_workbook
from datetime import date
# Connect to the database
cnxn = pyodbc.connect(
    'DRIVER={ODBC Driver 17 for SQL Server};SERVER=ES00VPADOSQL180,51433;DATABASE=SEO_MART;UID=your_username;PWD=your_password')
cursor = cnxn.cursor()
# SQL queries
qry_report = "SELECT * from [SEO_MART].[dbo].[RPT_RSProvisioning] where SUBSTRING(EnrolledDBN,1,2)<'84'"
comp_report = "SELECT * from [SEO_MART].[dbo].[RPT_RSCompliance] where SchoolDistrict<84"
comp_report2 = "SELECT * from [SEO_MART].[dbo].[RPT_RSCompliancebyGroup]"
comp_report3 = "SELECT * from [SEO_MART].[dbo].[RPT_Locations]"
# Read data from SQL Server
report_RS = pd.read_sql_query(qry_report, cnxn)
report_citywide = pd.read_sql_query(comp_report, cnxn)
report_citywide2 = pd.read_sql_query(comp_report2, cnxn)
report_lcgms = pd.read_sql_query(comp_report3, cnxn)
# Close the connection
cnxn.close()
# Convert datetimes to dates
report_RS_p2 = report_RS.copy()
report_RS_p2 = report_RS_p2.select_dtypes(
    include=[np.datetime64]).apply(pd.Series.dt.date)
# Replace the original table
report_RS = report_RS_p2
# Compute report_RS_Temp1
report_RS["STUDENTNAME"] = report_RS["LastName"] + "," + report_RS["FirstName"]
report_RS["Attendrate"] = (report_RS["AttendRate"]
                           * 100).round(0).astype(str) + "%"
report_RS["district"] = report_RS["EnrolledDBN"].str[:2]
report_RS_Temp1 = report_RS[[
    "StudentID", "STUDENTNAME", "Attendrate", "ServiceType", "RecommendedGroupSizeNumeric",
    "RecommendedFrequencyNumeric", "RecommendedDurationNumeric", "RSMandateLanguage", "EnrolledDBN", "GradeLevel",
    "BirthDate", "EffectiveOutcomeDate", "RecentAuthorizationDate", "PhysicalLocation", "PhysicalLocationName",
    "PhysicalLocationZipCode", "MandateType", "FirstEncounterDate", "PAFirstPartialAttendDate",
    "SESISFirstPartialEncounterDate", "TotalPartialEncounters", "SESISLastPartialEncounterDate", "district",
    "FirmName", "EncounterProvider", "ProcessedDate"
]].copy()

# Define SQL queries
report_RS_Temp_query = """
    SELECT *
    FROM report_RS_Temp1
    WHERE district < 84
"""
report_RS_comp_query = """
    SELECT *
    FROM report_citywide
"""
report_RS_comp2_query = """
    SELECT a.*, b.SuperintendentDistrict, a.SuperintendentName
    FROM report_citywide AS a
    LEFT JOIN report_lcgms AS b
    ON a.EnrolledDBN = b.SchoolDBN
"""
report_RS_dbn_query = """
    SELECT *
    FROM report_RS_comp2
"""
report_RS_dst_query = """
    SELECT *
    FROM report_citywide2
    WHERE ReportGroupDesc = 'SchoolDistrict'
    ORDER BY SchoolDistrict
"""
report_RS_sup_query = """
    SELECT *
    FROM report_citywide2
    WHERE ReportGroupDesc = 'Superintendent'
    ORDER BY SuperintendentName
"""
report_asofdt_query = """
    SELECT ProcessedDate
    FROM report_RS
    LIMIT 1
"""
# Read data from SQL Server
report_RS_Temp = pd.read_sql_query(report_RS_Temp_query, cnxn)
report_RS_comp = pd.read_sql_query(report_RS_comp_query, cnxn)
report_RS_comp2 = pd.read_sql_query(report_RS_comp2_query, cnxn)
report_RS_dbn = pd.read_sql_query(report_RS_dbn_query, cnxn)
report_RS_dst = pd.read_sql_query(report_RS_dst_query, cnxn)
report_RS_sup = pd.read_sql_query(report_RS_sup_query, cnxn)
report_asofdt = pd.read_sql_query(report_asofdt_query, cnxn)
# Close the connection
cnxn.close()
# Current date
dt = date.today()

################################################################
start = pd.Timestamp.now()
dt = date.today().strftime("%Y%m%d")
pth2 = os.path.join("R:/SEO Analytics/Share/Related Services/", dt)
os.makedirs(pth2, exist_ok=True)
mycomp1 = report_RS_dbn
mycomp2 = report_RS_dst
mycomp3 = report_RS_sup
asofdt = report_asofdt
wb = load_workbook("C:/Template/RS_Compliance_new.xlsx")
sheet = wb["RS Supt Citywide Summary"]
sheet.cell(row=1, column=3, value=asofdt)
for idx, row in mycomp1.iterrows():
    for jdx, value in enumerate(row):
        sheet.cell(row=8+idx, column=1+jdx, value=value)
        sheet = wb["RS District Summary"]
sheet.cell(row=1, column=3, value=asofdt)
for idx, row in mycomp2.iterrows():
    for jdx, value in enumerate(row):
        sheet.cell(row=8+idx, column=1+jdx, value=value)
        sheet = wb["RS Superintendent Summary"]
sheet.cell(row=1, column=3, value=asofdt)
for idx, row in mycomp3.iterrows():
    for jdx, value in enumerate(row):
        sheet.cell(row=8+idx, column=1+jdx, value=value)
        pth1 = os.path.join(pth2, f"RS Compliance Report_{dt}.xlsx")
wb.save(pth1)
end = pd.Timestamp.now()
print(f"Start time: {start}, End time: {end}")
